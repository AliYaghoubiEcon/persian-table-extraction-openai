import os
import json
import time
import base64
import pandas as pd
from tqdm import tqdm
from openpyxl import Workbook
from openai import OpenAI

# =========================================================
# OpenAI configuration
# =========================================================
# IMPORTANT:
# Set your API key as an environment variable before running:
# export OPENAI_API_KEY="your_api_key_here"
# or on Windows:
# setx OPENAI_API_KEY "your_api_key_here"

MODEL = "gpt-5.2"

SYSTEM_PROMPT = """
You are an expert table digitizer.

Given ONE image that contains a table (any language, any layout),
extract the full table structure.

The title of the table (usually in red) contains:
- province → column "استان"
- university name → column "نام دانشگاه"

Be careful with column "توضیحات":
it may contain multiple lines.

Return STRICT JSON ONLY with this schema:
{
  "sheet_name": "<short name>",
  "columns": ["col1","col2",...],
  "rows": [
    ["r1c1","r1c2",...]
  ]
}

Rules:
- Preserve text exactly as seen.
- Keep column order left-to-right.
- Use "" for empty cells.
- If unreadable, write "??".
"""

client = OpenAI()

# =========================================================
# Paths
# =========================================================
INPUT_DIR = "data/raw/images"
OUTPUT_DIR = "data/processed"
PER_IMAGE_DIR = os.path.join(OUTPUT_DIR, "per_image")

os.makedirs(PER_IMAGE_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

PROGRESS_FILE = os.path.join(OUTPUT_DIR, "progress.json")
ERROR_LOG = os.path.join(OUTPUT_DIR, "errors.csv")
MERGED_EXCEL = os.path.join(OUTPUT_DIR, "merged_all.xlsx")
SHEETS_EXCEL = os.path.join(OUTPUT_DIR, "all_images_sheets.xlsx")

# =========================================================
# Resume helpers
# =========================================================
def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
            return set(json.load(f))
    return set()

def save_progress(done):
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(list(done), f, ensure_ascii=False, indent=2)

# =========================================================
# Image utilities
# =========================================================
def img_to_data_url(path):
    """Convert image to base64 data URL for OpenAI vision models."""
    with open(path, "rb") as f:
        b64 = base64.b64encode(f.read()).decode("utf-8")
    return f"data:image/jpeg;base64,{b64}"

# =========================================================
# LLM-based table extraction
# =========================================================
def extract_table(image_path):
    """Send image to OpenAI and parse structured table output."""
    response = client.responses.create(
        model=MODEL,
        input=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {
                "role": "user",
                "content": [
                    {"type": "input_text", "text": "Extract the table from this image."},
                    {
                        "type": "input_image",
                        "image_url": img_to_data_url(image_path),
                        "detail": "high",
                    },
                ],
            },
        ],
    )
    return json.loads(response.output_text.strip())

# =========================================================
# Excel helpers
# =========================================================
def save_excel(path, payload):
    """Save a single extracted table to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = payload.get("sheet_name", "Sheet1")[:31]

    columns = payload["columns"]
    rows = payload["rows"]

    ws.append(columns)
    for r in rows:
        r = list(r) + [""] * (len(columns) - len(r))
        ws.append(r[:len(columns)])

    wb.save(path)

# =========================================================
# Main processing loop
# =========================================================
processed = load_progress()
images = sorted(
    f for f in os.listdir(INPUT_DIR)
    if f.lower().endswith((".jpg", ".jpeg", ".png"))
)

all_tables = []
sheet_payloads = {}

for img in tqdm(images):
    if img in processed:
        continue

    img_path = os.path.join(INPUT_DIR, img)
    out_xlsx = os.path.join(PER_IMAGE_DIR, os.path.splitext(img)[0] + ".xlsx")

    try:
        payload = extract_table(img_path)
        save_excel(out_xlsx, payload)

        df = pd.DataFrame(payload["rows"], columns=payload["columns"])
        df["source_image"] = img
        all_tables.append(df)

        sheet_payloads[img] = payload
        processed.add(img)
        save_progress(processed)

        time.sleep(20)  # Rate limiting

    except Exception as e:
        pd.DataFrame([[img, str(e)]], columns=["image", "error"]).to_csv(
            ERROR_LOG,
            mode="a",
            index=False,
            header=not os.path.exists(ERROR_LOG),
        )

# =========================================================
# Final outputs
# =========================================================
if all_tables:
    merged_df = pd.concat(all_tables, ignore_index=True)
    merged_df.to_excel(MERGED_EXCEL, index=False)

wb = Workbook()
wb.remove(wb.active)

for img, payload in sheet_payloads.items():
    ws = wb.create_sheet(os.path.splitext(img)[0][:31])
    ws.append(payload["columns"])
    for r in payload["rows"]:
        r = list(r) + [""] * (len(payload["columns"]) - len(r))
        ws.append(r[:len(payload["columns"])])

wb.save(SHEETS_EXCEL)

print("✅ All done!")
print("Output folder:", OUTPUT_DIR)
